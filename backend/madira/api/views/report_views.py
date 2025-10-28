import csv
import io
import time
from datetime import datetime, timedelta
from decimal import Decimal
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Sum, Count, Q, F
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.core.cache import cache

from ..models import (
    Order, Input, Output, Client, Supplier, Product, 
    StockMovement, OrderOutput, User
)


class SystemBenchmark:
    """
    Benchmark the actual system performance for accurate time estimation.
    Results are cached for 1 hour to avoid repeated benchmarking.
    """
    CACHE_KEY = 'system_benchmark_results'
    CACHE_TIMEOUT = 3600  # 1 hour
    
    @classmethod
    def get_benchmarks(cls):
        """Get cached benchmarks or run new benchmark test"""
        cached = cache.get(cls.CACHE_KEY)
        if cached:
            return cached
        
        # Run benchmark
        benchmarks = cls._run_benchmark()
        cache.set(cls.CACHE_KEY, benchmarks, cls.CACHE_TIMEOUT)
        return benchmarks
    
    @classmethod
    def _run_benchmark(cls):
        """Run actual performance tests on this hardware"""
        print("🔧 Running system benchmark...")
        
        # Test 1: Database query performance
        db_start = time.time()
        try:
            # Query a sample of data with joins
            Order.objects.select_related('client').all()[:100].count()
            Input.objects.select_related('created_by', 'order').all()[:100].count()
            Output.objects.select_related('created_by', 'source_input').all()[:100].count()
        except:
            pass
        db_time = time.time() - db_start
        db_query_per_1k = max(db_time * 10, 0.1)  # Extrapolate to 1K rows, minimum 0.1s
        
        # Test 2: Excel creation and writing
        excel_start = time.time()
        wb = Workbook()
        ws = wb.active
        
        # Write 100 rows × 10 columns
        for i in range(100):
            ws.append([f"Data {i}", i, i*2, i*3, "Text", "Value", i*10, i*20, "Info", "Notes"])
        
        excel_time = time.time() - excel_start
        excel_write_per_1k = max(excel_time * 10, 0.15)  # Extrapolate to 1K rows
        
        # Test 3: Styling performance
        style_start = time.time()
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Style header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        # Style 100 data rows
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=101), start=1):
            for cell in row:
                cell.border = border
                if idx % 2 == 0:
                    cell.fill = alt_fill
        
        style_time = time.time() - style_start
        style_per_1k = max(style_time * 10, 0.2)  # Extrapolate to 1K rows
        
        # Test 4: Save to memory
        save_start = time.time()
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        file_size_mb = len(output.getvalue()) / (1024 * 1024)
        save_time = time.time() - save_start
        
        benchmarks = {
            'db_query_per_1k': round(db_query_per_1k, 3),
            'excel_write_per_1k': round(excel_write_per_1k, 3),
            'style_per_1k': round(style_per_1k, 3),
            'save_base': round(save_time, 3),
            'save_per_mb': round(save_time / max(file_size_mb, 0.01), 3),
            'tested_at': datetime.now().isoformat(),
            'hardware_info': {
                'python_version': f"{__import__('sys').version_info.major}.{__import__('sys').version_info.minor}",
                'platform': __import__('platform').system(),
            }
        }
        
        print(f"✅ Benchmark complete: DB={db_query_per_1k:.3f}s, Excel={excel_write_per_1k:.3f}s, Style={style_per_1k:.3f}s per 1K rows")
        
        return benchmarks


class ReportEstimateView(APIView):
    """
    Estimate report generation time before downloading.
    GET /api/reports/estimate/
    
    Uses actual hardware benchmarks for accurate time prediction.
    """
    permission_classes = [IsAuthenticated]
    
    def get_date_range(self, report_type, custom_start=None, custom_end=None):
        """Calculate date range based on report type"""
        now = timezone.now()
        
        if custom_start and custom_end:
            start_date = datetime.fromisoformat(custom_start.replace('Z', '+00:00'))
            end_date = datetime.fromisoformat(custom_end.replace('Z', '+00:00'))
        elif report_type == 'daily':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        elif report_type == 'weekly':
            start_date = now - timedelta(days=7)
            end_date = now
        elif report_type == 'monthly':
            start_date = now - timedelta(days=30)
            end_date = now
        elif report_type == 'all':
            start_date = datetime(2000, 1, 1, tzinfo=now.tzinfo)
            end_date = now
        else:
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        
        return start_date, end_date
    
    def get(self, request):
        report_type = request.query_params.get('type', 'monthly')
        custom_start = request.query_params.get('start_date')
        custom_end = request.query_params.get('end_date')
        force_benchmark = request.query_params.get('force_benchmark', 'false').lower() == 'true'
        
        # Get filters
        client_id = request.query_params.get('client_id')
        order_id = request.query_params.get('order_id')
        supplier_id = request.query_params.get('supplier_id')
        product_id = request.query_params.get('product_id')
        status_filter = request.query_params.get('status')
        include_relations = request.query_params.get('include_relations', 'true').lower() == 'true'
        
        start_date, end_date = self.get_date_range(report_type, custom_start, custom_end)
        
        # ============================================
        # GET HARDWARE BENCHMARKS
        # ============================================
        if force_benchmark:
            cache.delete(SystemBenchmark.CACHE_KEY)
        
        benchmarks = SystemBenchmark.get_benchmarks()
        
        # ============================================
        # ACCURATE DATABASE COUNTING WITH TIMING
        # ============================================
        count_start = time.time()
        
        # Count transactional data (filtered by date range)
        orders_query = Order.objects.filter(order_date__range=[start_date, end_date])
        if client_id:
            orders_query = orders_query.filter(client_id=client_id)
        if order_id:
            orders_query = orders_query.filter(id=order_id)
        if status_filter:
            orders_query = orders_query.filter(status=status_filter)
        orders_count = orders_query.count()
        
        inputs_query = Input.objects.filter(date__range=[start_date, end_date])
        if client_id:
            inputs_query = inputs_query.filter(order__client_id=client_id)
        if order_id:
            inputs_query = inputs_query.filter(order_id=order_id)
        inputs_count = inputs_query.count()
        
        outputs_query = Output.objects.filter(date__range=[start_date, end_date])
        if order_id:
            outputs_query = outputs_query.filter(order_id=order_id)
        if supplier_id:
            outputs_query = outputs_query.filter(supplier_id=supplier_id)
        if product_id:
            outputs_query = outputs_query.filter(product_id=product_id)
        outputs_count = outputs_query.count()
        
        order_outputs_query = OrderOutput.objects.filter(created_at__range=[start_date, end_date])
        if order_id:
            order_outputs_query = order_outputs_query.filter(order_id=order_id)
        if client_id:
            order_outputs_query = order_outputs_query.filter(order__client_id=client_id)
        order_outputs_count = order_outputs_query.count()
        
        movements_query = StockMovement.objects.filter(date__range=[start_date, end_date])
        if order_id:
            movements_query = movements_query.filter(order_id=order_id)
        if product_id:
            movements_query = movements_query.filter(product_id=product_id)
        if client_id:
            movements_query = movements_query.filter(order__client_id=client_id)
        movements_count = movements_query.count()
        
        # Count master data (NOT filtered by date)
        clients_query = Client.objects.all()
        if client_id:
            clients_query = clients_query.filter(id=client_id)
        clients_count = clients_query.count()
        
        suppliers_query = Supplier.objects.all()
        if supplier_id:
            suppliers_query = suppliers_query.filter(id=supplier_id)
        suppliers_count = suppliers_query.count()
        
        products_query = Product.objects.all()
        if product_id:
            products_query = products_query.filter(id=product_id)
        products_count = products_query.count()
        
        users_count = User.objects.count()
        
        # Measure actual count query time
        count_query_time = time.time() - count_start
        
        # ============================================
        # CALCULATE HISTORY SHEETS
        # ============================================
        history_sheets_count = 0
        history_rows = 0
        history_sheets_breakdown = {}
        
        if include_relations:
            if client_id:
                history_sheets_count += 1
                client_orders = orders_query.count()
                client_payments = inputs_query.count()
                client_history_rows = 3 + 11 + 6 + 2 + client_orders + 1 + 2 + client_payments
                history_rows += client_history_rows
                history_sheets_breakdown['client_history'] = client_history_rows
                
            if order_id:
                history_sheets_count += 1
                order_payments = inputs_query.count()
                order_expenses = order_outputs_query.count()
                order_movements = movements_query.count()
                order_history_rows = 3 + 15 + 1 + 2 + order_payments + 1 + 2 + order_expenses + 1 + 2 + order_movements
                history_rows += order_history_rows
                history_sheets_breakdown['order_history'] = order_history_rows
                
            if supplier_id:
                history_sheets_count += 1
                supplier_outputs = outputs_query.count()
                supplier_history_rows = 3 + 9 + 1 + 3 + 1 + 2 + supplier_outputs
                history_rows += supplier_history_rows
                history_sheets_breakdown['supplier_history'] = supplier_history_rows
                
            if product_id:
                history_sheets_count += 1
                product_movements = movements_query.count()
                product_history_rows = 3 + 10 + 1 + 4 + 1 + 2 + product_movements
                history_rows += product_history_rows
        
        # ============================================
        # CALCULATE EXACT EXCEL ROWS
        # ============================================
        SHEET_OVERHEAD = 4
        
        transactional_data_rows = orders_count + inputs_count + outputs_count + order_outputs_count + movements_count
        master_data_rows = clients_count + suppliers_count + products_count + users_count
        total_data_rows = transactional_data_rows + master_data_rows
        
        base_sheets = 9
        total_sheets = base_sheets + history_sheets_count
        total_overhead_rows = total_sheets * SHEET_OVERHEAD
        total_excel_rows = total_data_rows + total_overhead_rows + history_rows
        
        # ============================================
        # HARDWARE-BASED TIME ESTIMATION
        # ============================================
        
        # Use actual benchmarked performance from this machine
        DB_QUERY_BASE = 0.3
        DB_QUERY_PER_1K = benchmarks['db_query_per_1k']
        
        EXCEL_CREATE_WORKBOOK = 0.2
        EXCEL_CREATE_SHEET = 0.05
        EXCEL_WRITE_PER_1K = benchmarks['excel_write_per_1k']
        
        STYLE_HEADER_PER_SHEET = 0.08
        STYLE_DATA_PER_1K = benchmarks['style_per_1k']
        
        AUTOFILTER_PER_SHEET = 0.1
        COLUMN_WIDTH_PER_SHEET = 0.12
        
        FILE_SAVE_BASE = benchmarks['save_base']
        FILE_SAVE_PER_MB = benchmarks['save_per_mb']
        HTTP_RESPONSE = 0.15
        
        # Calculate each component
        db_query_time = DB_QUERY_BASE + (total_data_rows / 1000.0) * DB_QUERY_PER_1K
        
        excel_creation_time = EXCEL_CREATE_WORKBOOK + (total_sheets * EXCEL_CREATE_SHEET)
        excel_write_time = (total_excel_rows / 1000.0) * EXCEL_WRITE_PER_1K
        
        styling_header_time = total_sheets * STYLE_HEADER_PER_SHEET
        styling_data_time = (total_excel_rows / 1000.0) * STYLE_DATA_PER_1K
        
        filter_time = total_sheets * AUTOFILTER_PER_SHEET
        column_time = total_sheets * COLUMN_WIDTH_PER_SHEET
        
        # Estimate file size
        BASE_WORKBOOK = 8000
        BYTES_PER_DATA_ROW = 280
        BYTES_PER_OVERHEAD_ROW = 180
        BYTES_PER_SHEET = 2200
        STYLING_MULTIPLIER = 1.35
        
        raw_size_bytes = (
            BASE_WORKBOOK +
            (total_sheets * BYTES_PER_SHEET) +
            (total_data_rows * BYTES_PER_DATA_ROW) +
            (total_overhead_rows * BYTES_PER_OVERHEAD_ROW) +
            (history_rows * BYTES_PER_DATA_ROW)
        )
        estimated_size_bytes = int(raw_size_bytes * STYLING_MULTIPLIER)
        estimated_size_mb = estimated_size_bytes / (1024 * 1024)
        
        file_save_time = FILE_SAVE_BASE + (estimated_size_mb * FILE_SAVE_PER_MB)
        http_response_time = HTTP_RESPONSE
        
        # Total base time
        base_time_estimate = (
            db_query_time +
            excel_creation_time +
            excel_write_time +
            styling_header_time +
            styling_data_time +
            filter_time +
            column_time +
            file_save_time +
            http_response_time
        )
        
        # ============================================
        # DYNAMIC ADJUSTMENTS BASED ON ACTUAL PERFORMANCE
        # ============================================
        
        # Use actual count query time to calibrate
        if count_query_time > 3.0:
            db_slowness_factor = 1.0 + (count_query_time / 3.0)
            base_time_estimate *= min(db_slowness_factor, 4.0)
        elif count_query_time > 1.5:
            base_time_estimate *= 1.5
        elif count_query_time > 0.8:
            base_time_estimate *= 1.25
        
        # Large dataset overhead (memory, GC, CPU)
        if total_data_rows > 200000:
            base_time_estimate *= 2.2  # Reduced from 3.0
        elif total_data_rows > 100000:
            base_time_estimate *= 1.8  # Reduced from 2.5
        elif total_data_rows > 50000:
            base_time_estimate *= 1.5  # Reduced from 2.0
        elif total_data_rows > 20000:
            base_time_estimate *= 1.3  # Reduced from 1.5
        elif total_data_rows > 10000:
            base_time_estimate *= 1.15  # Reduced from 1.25
        
        # Many sheets overhead
        if total_sheets > 15:
            base_time_estimate *= 1.4
        elif total_sheets > 10:
            base_time_estimate *= 1.2
        
        total_time_estimate = base_time_estimate
        
        # Add confidence margin based on variance
        confidence_margin = total_time_estimate * 0.15  # 15% margin
        min_time = max(total_time_estimate - confidence_margin, total_time_estimate * 0.7)
        max_time = total_time_estimate + confidence_margin
        
        # ============================================
        # RECOMMENDATIONS & WARNINGS
        # ============================================
        
        if total_time_estimate < 3:
            recommendation = "very_fast"
            message = "⚡ This report will generate very quickly (under 3 seconds)."
        elif total_time_estimate < 8:
            recommendation = "fast"
            message = "✅ This report will generate quickly (3-8 seconds)."
        elif total_time_estimate < 20:
            recommendation = "moderate"
            message = "⏱️ This report will take 8-20 seconds to generate."
        elif total_time_estimate < 60:
            recommendation = "slow"
            message = "⚠️ This report will take 20-60 seconds. Please be patient."
        elif total_time_estimate < 180:
            recommendation = "very_slow"
            message = f"⏳ This report will take 1-3 minutes. Consider filtering the data."
        else:
            recommendation = "extremely_slow"
            minutes = int(total_time_estimate / 60)
            message = f"🐌 This report will take over 3 minutes (~{minutes} min). Strongly recommend using filters."
        
        # ============================================
        # RESPONSE
        # ============================================
        
        return Response({
            'success': True,
            'time_estimate': {
                'total_seconds': round(total_time_estimate, 2),
                'min_seconds': round(min_time, 2),
                'max_seconds': round(max_time, 2),
                'readable': self._format_time(total_time_estimate),
                'range_readable': f"{self._format_time(min_time)} - {self._format_time(max_time)}",
                'breakdown': {
                    'database_query': round(db_query_time, 2),
                    'excel_creation': round(excel_creation_time, 2),
                    'excel_writing': round(excel_write_time, 2),
                    'header_styling': round(styling_header_time, 2),
                    'data_styling': round(styling_data_time, 2),
                    'autofilters': round(filter_time, 2),
                    'column_widths': round(column_time, 2),
                    'file_saving': round(file_save_time, 2),
                    'http_response': round(http_response_time, 2)
                }
            },
            'recommendation': recommendation,
            'message': message
        })
    
    def _format_time(self, seconds):
        """Format seconds into human-readable time"""
        if seconds < 60:
            return f"{int(seconds)} seconds"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            secs = int(seconds % 60)
            return f"{minutes} min {secs} sec"
        else:
            hours = int(seconds / 3600)
            minutes = int((seconds % 3600) / 60)
            return f"{hours} hr {minutes} min"
    
    def _format_size(self, size_bytes):
        """Format bytes into human-readable size"""
        if size_bytes < 1024:
            return f"{size_bytes} bytes"
        elif size_bytes < 1024 * 1024:
            return f"{round(size_bytes / 1024, 2)} KB"
        elif size_bytes < 1024 * 1024 * 1024:
            return f"{round(size_bytes / (1024 * 1024), 2)} MB"
        else:
            return f"{round(size_bytes / (1024 * 1024 * 1024), 2)} GB"
    
    def _get_suggestions(self, total_rows, report_type, recommendation, filters):
        """Provide suggestions to optimize report generation"""
        suggestions = []
        
        if recommendation in ['slow', 'very_slow', 'extremely_slow', 'too_large']:
            suggestions.append("Consider using a shorter time period (daily or weekly instead of monthly/all)")
            if not filters.get('client_id'):
                suggestions.append("Filter by specific client to reduce data volume")
            if not filters.get('status'):
                suggestions.append("Filter by order status (e.g., only 'completed' orders)")
            
        if total_rows > 50000:
            suggestions.append("Use filters to focus on specific data you need")
            suggestions.append("Generate separate reports for different time periods")
            
        if report_type == 'all':
            suggestions.append("'All time' reports can be very large. Consider monthly or quarterly reports instead")
        
        if not suggestions:
            suggestions.append("Your report is optimized and ready to download!")
        
        return suggestions


class ComprehensiveReportView(APIView):
    """
    Detailed backup-style Excel report with ALL business data and filtering.
    GET /api/reports/download/
    
    Query Parameters:
        - type: 'daily', 'weekly', 'monthly', 'all' (default: 'monthly')
        - start_date: Custom start date (ISO format)
        - end_date: Custom end date (ISO format)
        - client_id: Filter by specific client
        - order_id: Filter by specific order
        - supplier_id: Filter by specific supplier
        - product_id: Filter by specific product
        - status: Filter orders by status
        - include_relations: 'true' to include all relationship data (default: 'true')
    
    Returns: Excel file with detailed sheets containing complete database records with relationships
    """
    permission_classes = [IsAuthenticated]

    def get_date_range(self, report_type, custom_start=None, custom_end=None):
        """Calculate date range based on report type"""
        now = timezone.now()
        
        if custom_start and custom_end:
            start_date = datetime.fromisoformat(custom_start.replace('Z', '+00:00'))
            end_date = datetime.fromisoformat(custom_end.replace('Z', '+00:00'))
        elif report_type == 'daily':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        elif report_type == 'weekly':
            start_date = now - timedelta(days=7)
            end_date = now
        elif report_type == 'monthly':
            start_date = now - timedelta(days=30)
            end_date = now
        elif report_type == 'all':
            start_date = datetime(2000, 1, 1, tzinfo=now.tzinfo)
            end_date = now
        else:
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        
        return start_date, end_date

    def apply_header_styling(self, ws, title, subtitle=""):
        """Apply professional header styling to worksheet"""
        title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        title_font = Font(bold=True, size=18, color="FFFFFF", name="Calibri")
        
        ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        if subtitle:
            ws.merge_cells(f'A2:{get_column_letter(ws.max_column)}2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = subtitle
            subtitle_cell.font = Font(italic=True, size=11, color="404040", name="Calibri")
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
            subtitle_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws.row_dimensions[2].height = 22

    def apply_table_styling(self, ws, header_row=4, enable_filters=True):
        """Apply professional table styling with Excel filters"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        
        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        border = Border(
            left=Side(style='thin', color='A6A6A6'),
            right=Side(style='thin', color='A6A6A6'),
            top=Side(style='thin', color='A6A6A6'),
            bottom=Side(style='thin', color='A6A6A6')
        )
        
        # Header row styling
        for cell in ws[header_row]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
        
        ws.row_dimensions[header_row].height = 30
        
        # Enable Excel AutoFilter on header row
        if enable_filters and ws.max_row > header_row:
            # Get the range for filters (from first column to last column of header row)
            filter_range = f'A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}'
            ws.auto_filter.ref = filter_range
        
        # Data rows with alternating colors
        for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row), start=1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                cell.font = Font(name="Calibri", size=10, color="000000")
                if idx % 2 == 0:
                    cell.fill = alt_fill
                else:
                    cell.fill = white_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 4, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

    def get(self, request):
        # Start timing the entire process
        generation_start_time = time.time()
        
        report_type = request.query_params.get('type', 'monthly')
        custom_start = request.query_params.get('start_date')
        custom_end = request.query_params.get('end_date')
        
        # Get filters
        client_id = request.query_params.get('client_id')
        order_id = request.query_params.get('order_id')
        supplier_id = request.query_params.get('supplier_id')
        product_id = request.query_params.get('product_id')
        status = request.query_params.get('status')
        include_relations = request.query_params.get('include_relations', 'true').lower() == 'true'
        
        filters = {
            'client_id': client_id,
            'order_id': order_id,
            'supplier_id': supplier_id,
            'product_id': product_id,
            'status': status,
            'include_relations': include_relations
        }
        
        start_date, end_date = self.get_date_range(report_type, custom_start, custom_end)
        
        # Print to console - Start
        print("="*80)
        print(f"📊 REPORT GENERATION STARTED")
        print(f"   Type: {report_type.upper()}")
        print(f"   Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        print(f"   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("="*80)
        
        # Generate the report
        response = self._export_detailed_excel(start_date, end_date, report_type, filters)
        
        # Calculate total generation time
        total_generation_time = time.time() - generation_start_time
        
        # Print to console - Complete
        print("="*80)
        print(f"✅ REPORT GENERATION COMPLETED")
        print(f"   Total Time: {total_generation_time:.2f} seconds ({self._format_time(total_generation_time)})")
        print(f"   File Ready for Download")
        print(f"   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("="*80)
        
        return response
    
    def _format_time(self, seconds):
        """Format seconds into human-readable time"""
        if seconds < 60:
            return f"{int(seconds)} seconds"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            secs = int(seconds % 60)
            return f"{minutes} min {secs} sec"
        else:
            hours = int(seconds / 3600)
            minutes = int((seconds % 3600) / 60)
            return f"{hours} hr {minutes} min"

    def _export_detailed_excel(self, start_date, end_date, report_type, filters):
        wb = Workbook()
        wb.remove(wb.active)
        
        # Detailed sheets with ALL data
        self._add_orders_detailed_sheet(wb, start_date, end_date, report_type, filters)
        self._add_inputs_detailed_sheet(wb, start_date, end_date, report_type, filters)
        self._add_outputs_detailed_sheet(wb, start_date, end_date, report_type, filters)
        self._add_order_outputs_detailed_sheet(wb, start_date, end_date, report_type, filters)
        self._add_stock_movements_detailed_sheet(wb, start_date, end_date, report_type, filters)
        
        if filters['include_relations']:
            # Add relationship sheets
            if filters['client_id']:
                self._add_client_history_sheet(wb, filters['client_id'], start_date, end_date)
            if filters['order_id']:
                self._add_order_history_sheet(wb, filters['order_id'])
            if filters['supplier_id']:
                self._add_supplier_history_sheet(wb, filters['supplier_id'], start_date, end_date)
            if filters['product_id']:
                self._add_product_history_sheet(wb, filters['product_id'], start_date, end_date)
        
        # Master data sheets
        self._add_clients_detailed_sheet(wb, filters)
        self._add_suppliers_detailed_sheet(wb, filters)
        self._add_products_detailed_sheet(wb, filters)
        self._add_users_detailed_sheet(wb)
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filter_desc = []
        if filters['client_id']:
            filter_desc.append(f"Client_{filters['client_id']}")
        if filters['order_id']:
            filter_desc.append(f"Order_{filters['order_id']}")
        if filters['status']:
            filter_desc.append(f"{filters['status']}")
        
        filter_str = "_".join(filter_desc) if filter_desc else ""
        filename = f"Detailed_Report_{report_type.upper()}_{filter_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    def _add_orders_detailed_sheet(self, wb, start_date, end_date, report_type, filters):
        ws = wb.create_sheet("Orders - Detailed", 0)
        
        orders = Order.objects.filter(
            order_date__range=[start_date, end_date]
        ).select_related('client')
        
        # Apply filters
        if filters['client_id']:
            orders = orders.filter(client_id=filters['client_id'])
        if filters['order_id']:
            orders = orders.filter(id=filters['order_id'])
        if filters['status']:
            orders = orders.filter(status=filters['status'])
        
        orders = orders.order_by('-order_date')
        
        title = f"DETAILED ORDERS REPORT - {report_type.upper()}"
        subtitle = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')} | Total Records: {orders.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Order Number', 'Client Name', 'Client Phone', 
            'Order Date', 'Delivery Date', 'Status', 
            'Total Amount (DA)', 'Paid Amount (DA)', 'Remaining (DA)',
            'Total Expenses (DA)', 'Net Benefit (DA)', 'Benefit %',
            'Payment Status', 'Total Payments Received', 'Description'
        ]
        ws.append(headers)
        
        for order in orders:
            benefit_pct = (float(order.total_benefit) / float(order.total_amount) * 100) if order.total_amount > 0 else 0
            payments_count = order.payments.count()
            
            ws.append([
                order.order_number,
                order.client.name,
                order.client.phone,
                order.order_date.strftime('%Y-%m-%d %H:%M'),
                order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else 'Not Set',
                order.get_status_display(),
                float(order.total_amount),
                float(order.paid_amount),
                float(order.remaining_amount),
                float(order.total_expenses),
                float(order.total_benefit),
                f"{benefit_pct:.2f}%",
                'Fully Paid' if order.is_fully_paid else 'Pending Payment',
                payments_count,
                order.description
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
    
    def _add_inputs_detailed_sheet(self, wb, start_date, end_date, report_type, filters):
        ws = wb.create_sheet("Inputs - Detailed")
        
        inputs = Input.objects.filter(
            date__range=[start_date, end_date]
        ).select_related('created_by', 'order', 'order__client')
        
        inputs = inputs.order_by('-date')
        
        title = f"DETAILED INPUTS (INCOME) REPORT - {report_type.upper()}"
        subtitle = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')} | Total Records: {inputs.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Reference Number', 'Type', 'Amount (DA)', 'Date & Time',
            'Client Name', 'Client Phone', 'Order Number', 
            'Total Spent (DA)', 'Remaining Balance (DA)',
            'Created By', 'Description'
        ]
        ws.append(headers)
        
        for inp in inputs:
            # Calculate total spent from this input (sum of all related outputs)
            total_spent = inp.outputs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            remaining = inp.amount - total_spent
            
            ws.append([
                inp.reference,
                inp.get_type_display(),
                float(inp.amount),
                inp.date.strftime('%Y-%m-%d %H:%M'),
                inp.order.client.name if inp.order else '',
                inp.order.client.phone if inp.order else '',
                inp.order.order_number if inp.order else '',
                float(total_spent),
                float(remaining),
                inp.created_by.full_name or inp.created_by.username,
                inp.description
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
    
    def _add_outputs_detailed_sheet(self, wb, start_date, end_date, report_type, filters):
        ws = wb.create_sheet("Outputs - Detailed")
        
        outputs = Output.objects.filter(
            date__range=[start_date, end_date]
        ).select_related('created_by', 'source_input', 'order', 'supplier', 'product')
        
        # Apply filters
        if filters['order_id']:
            outputs = outputs.filter(order_id=filters['order_id'])
        if filters['supplier_id']:
            outputs = outputs.filter(supplier_id=filters['supplier_id'])
        if filters['product_id']:
            outputs = outputs.filter(product_id=filters['product_id'])
        
        outputs = outputs.order_by('-date')
        
        title = f"DETAILED OUTPUTS (EXPENSES) REPORT - {report_type.upper()}"
        subtitle = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')} | Total Records: {outputs.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Reference Number', 'Type', 'Amount (DA)', 'Date & Time',
            'Source Input Reference', 'Source Amount (DA)',
            'Order Number', 'Order Client',
            'Supplier Name', 'Product Name', 
            'Created By', 'Description'
        ]
        ws.append(headers)
        
        for out in outputs:
            ws.append([
                out.reference,
                out.get_type_display(),
                float(out.amount),
                out.date.strftime('%Y-%m-%d %H:%M'),
                out.source_input.reference,
                float(out.source_input.amount),
                out.order.order_number if out.order else '',
                out.order.client.name if out.order else '',
                out.supplier.name if out.supplier else '',
                out.product.name if out.product else '',
                out.created_by.full_name or out.created_by.username,
                out.description
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
    
    def _add_order_outputs_detailed_sheet(self, wb, start_date, end_date, report_type, filters):
        ws = wb.create_sheet("Order Expenses - Detailed")
        
        order_outputs = OrderOutput.objects.filter(
            created_at__range=[start_date, end_date]
        ).select_related('order', 'order__client', 'created_by_output', 'created_by_stock_movement')
        
        # Apply filters
        if filters['order_id']:
            order_outputs = order_outputs.filter(order_id=filters['order_id'])
        if filters['client_id']:
            order_outputs = order_outputs.filter(order__client_id=filters['client_id'])
        
        order_outputs = order_outputs.order_by('-created_at')
        
        title = f"DETAILED ORDER EXPENSES REPORT - {report_type.upper()}"
        subtitle = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')} | Total Records: {order_outputs.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Order Number', 'Client Name', 'Expense Type', 'Amount (DA)', 
            'Expense Reference', 'Product/Item', 
            'Date', 'Description'
        ]
        ws.append(headers)
        
        for oo in order_outputs:
            ws.append([
                oo.order.order_number,
                oo.order.client.name,
                oo.get_type_display(),
                float(oo.amount),
                oo.created_by_output.reference if oo.created_by_output else '',
                oo.created_by_stock_movement.product.name if oo.created_by_stock_movement else '',
                oo.created_at.strftime('%Y-%m-%d %H:%M'),
                oo.description
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
    
    def _add_stock_movements_detailed_sheet(self, wb, start_date, end_date, report_type, filters):
        ws = wb.create_sheet("Stock Movements - Detailed")
        
        movements = StockMovement.objects.filter(
            date__range=[start_date, end_date]
        ).select_related('product', 'order', 'order__client', 'created_by', 'created_by_output')
        
        # Apply filters
        if filters['order_id']:
            movements = movements.filter(order_id=filters['order_id'])
        if filters['product_id']:
            movements = movements.filter(product_id=filters['product_id'])
        if filters['client_id']:
            movements = movements.filter(order__client_id=filters['client_id'])
        
        movements = movements.order_by('-date')
        
        title = f"DETAILED STOCK MOVEMENTS REPORT - {report_type.upper()}"
        subtitle = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')} | Total Records: {movements.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Product Reference', 'Product Name', 'Movement Type',
            'Quantity', 'Unit', 'Price per Unit (DA)', 'Total Value (DA)',
            'Order Number', 'Client Name', 
            'Date & Time', 'Created By'
        ]
        ws.append(headers)
        
        for mov in movements:
            total_value = float(mov.quantity) * float(mov.price)
            ws.append([
                mov.product.reference,
                mov.product.name,
                mov.get_movement_type_display(),
                float(mov.quantity),
                mov.product.get_unit_display(),
                float(mov.price),
                total_value,
                mov.order.order_number if mov.order else '',
                mov.order.client.name if mov.order else '',
                mov.date.strftime('%Y-%m-%d %H:%M'),
                mov.created_by.full_name or mov.created_by.username if mov.created_by else ''
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
    
    def _add_client_history_sheet(self, wb, client_id, start_date, end_date):
        """Complete history for a specific client"""
        ws = wb.create_sheet(f"Client {client_id} - History")
        
        try:
            client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            return
        
        title = f"CLIENT COMPLETE HISTORY: {client.name}"
        subtitle = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        # Client Info
        ws.append(['CLIENT INFORMATION'])
        ws.append(['Field', 'Value'])
        ws.append(['Client ID', client.id])
        ws.append(['Name', client.name])
        ws.append(['Phone', client.phone])
        ws.append(['Address', client.address])
        ws.append(['Type', client.get_client_type_display()])
        ws.append(['Credit Balance', float(client.credit_balance)])
        ws.append(['Status', 'Active' if client.is_active else 'Inactive'])
        ws.append([])
        
        # Orders Summary
        orders = client.orders.filter(order_date__range=[start_date, end_date])
        total_orders_value = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        
        # Calculate total paid by summing all CLIENT_PAYMENT inputs for these orders
        total_paid = Input.objects.filter(
            order__in=orders,
            type=Input.Type.CLIENT_PAYMENT
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        ws.append(['ORDERS SUMMARY'])
        ws.append(['Metric', 'Value'])
        ws.append(['Total Orders', orders.count()])
        ws.append(['Total Orders Value (DA)', float(total_orders_value)])
        ws.append(['Total Paid (DA)', float(total_paid)])
        ws.append(['Outstanding (DA)', float(total_orders_value - total_paid)])
        ws.append([])
        
        # All Orders
        ws.append(['ALL ORDERS'])
        ws.append(['Order Number', 'Order Date', 'Status', 'Total Amount (DA)', 'Paid (DA)', 'Remaining (DA)', 'Expenses (DA)', 'Benefit (DA)'])
        
        for order in orders:
            ws.append([
                order.order_number,
                order.order_date.strftime('%Y-%m-%d'),
                order.get_status_display(),
                float(order.total_amount),
                float(order.paid_amount),
                float(order.remaining_amount),
                float(order.total_expenses),
                float(order.total_benefit)
            ])
        ws.append([])
        
        # All Payments (Inputs)
        inputs = Input.objects.filter(
            order__client=client,
            date__range=[start_date, end_date]
        ).order_by('-date')
        
        ws.append(['ALL PAYMENTS'])
        ws.append(['Reference', 'Date', 'Amount (DA)', 'Order Number', 'Type', 'Description'])
        
        for inp in inputs:
            ws.append([
                inp.reference,
                inp.date.strftime('%Y-%m-%d %H:%M'),
                float(inp.amount),
                inp.order.order_number if inp.order else '',
                inp.get_type_display(),
                inp.description[:100] if inp.description else ''
            ])
        
        self.apply_header_styling(ws, title, subtitle)
    
    def _add_order_history_sheet(self, wb, order_id):
        """Complete history for a specific order"""
        ws = wb.create_sheet(f"Order {order_id} - History")
        
        try:
            order = Order.objects.select_related('client').get(id=order_id)
        except Order.DoesNotExist:
            return
        
        title = f"ORDER COMPLETE HISTORY: {order.order_number}"
        subtitle = f"Client: {order.client.name} | Status: {order.get_status_display()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        # Order Info
        ws.append(['ORDER INFORMATION'])
        ws.append(['Field', 'Value'])
        ws.append(['Order ID', order.id])
        ws.append(['Order Number', order.order_number])
        ws.append(['Client', order.client.name])
        ws.append(['Order Date', order.order_date.strftime('%Y-%m-%d %H:%M')])
        ws.append(['Delivery Date', order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else 'Not Set'])
        ws.append(['Status', order.get_status_display()])
        ws.append(['Total Amount (DA)', float(order.total_amount)])
        ws.append(['Paid Amount (DA)', float(order.paid_amount)])
        ws.append(['Remaining (DA)', float(order.remaining_amount)])
        ws.append(['Total Expenses (DA)', float(order.total_expenses)])
        ws.append(['Net Benefit (DA)', float(order.total_benefit)])
        ws.append(['Description', order.description])
        ws.append([])
        
        # Payments for this order
        payments = order.payments.all().order_by('-date')
        ws.append(['PAYMENTS RECEIVED'])
        ws.append(['Reference', 'Date', 'Amount (DA)', 'Type', 'Created By', 'Description'])
        
        for payment in payments:
            ws.append([
                payment.reference,
                payment.date.strftime('%Y-%m-%d %H:%M'),
                float(payment.amount),
                payment.get_type_display(),
                payment.created_by.username,
                payment.description[:100] if payment.description else ''
            ])
        ws.append([])
        
        # Order Outputs (Expenses)
        order_outputs = order.order_outputs.all().order_by('-created_at')
        ws.append(['ORDER EXPENSES'])
        ws.append(['Type', 'Amount (DA)', 'Date', 'Output Reference', 'Product', 'Description'])
        
        for oo in order_outputs:
            ws.append([
                oo.get_type_display(),
                float(oo.amount),
                oo.created_at.strftime('%Y-%m-%d %H:%M'),
                oo.created_by_output.reference if oo.created_by_output else '',
                oo.created_by_stock_movement.product.name if oo.created_by_stock_movement else '',
                oo.description[:100] if oo.description else ''
            ])
        ws.append([])
        
        # Stock Movements for this order
        movements = order.stock_movements.all().order_by('-date')
        ws.append(['STOCK MOVEMENTS'])
        ws.append(['Product', 'Type', 'Quantity', 'Unit', 'Price (DA)', 'Total (DA)', 'Date'])
        
        for mov in movements:
            ws.append([
                mov.product.name,
                mov.get_movement_type_display(),
                float(mov.quantity),
                mov.product.get_unit_display(),
                float(mov.price),
                float(mov.quantity) * float(mov.price),
                mov.date.strftime('%Y-%m-%d %H:%M')
            ])
        
        self.apply_header_styling(ws, title, subtitle)
    
    def _add_supplier_history_sheet(self, wb, supplier_id, start_date, end_date):
        """Complete history for a specific supplier"""
        ws = wb.create_sheet(f"Supplier {supplier_id} - History")
        
        try:
            supplier = Supplier.objects.get(id=supplier_id)
        except Supplier.DoesNotExist:
            return
        
        title = f"SUPPLIER COMPLETE HISTORY: {supplier.name}"
        subtitle = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        # Supplier Info
        ws.append(['SUPPLIER INFORMATION'])
        ws.append(['Field', 'Value'])
        ws.append(['Supplier ID', supplier.id])
        ws.append(['Name', supplier.name])
        ws.append(['Phone', supplier.phone])
        ws.append(['Address', supplier.address])
        ws.append(['Status', 'Active' if supplier.is_active else 'Inactive'])
        ws.append([])
        
        # All Outputs for this supplier
        outputs = Output.objects.filter(
            supplier=supplier,
            date__range=[start_date, end_date]
        ).order_by('-date')
        
        total_expenses = outputs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        ws.append(['EXPENSES SUMMARY'])
        ws.append(['Total Transactions', outputs.count()])
        ws.append(['Total Amount (DA)', float(total_expenses)])
        ws.append([])
        
        ws.append(['ALL TRANSACTIONS'])
        ws.append(['Reference', 'Date', 'Type', 'Amount (DA)', 'Order', 'Product', 'Description'])
        
        for output in outputs:
            ws.append([
                output.reference,
                output.date.strftime('%Y-%m-%d %H:%M'),
                output.get_type_display(),
                float(output.amount),
                output.order.order_number if output.order else '',
                output.product.name if output.product else '',
                output.description[:100] if output.description else ''
            ])
        
        self.apply_header_styling(ws, title, subtitle)
    
    def _add_product_history_sheet(self, wb, product_id, start_date, end_date):
        """Complete history for a specific product"""
        ws = wb.create_sheet(f"Product {product_id} - History")
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return
        
        title = f"PRODUCT COMPLETE HISTORY: {product.name}"
        subtitle = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        # Product Info
        ws.append(['PRODUCT INFORMATION'])
        ws.append(['Field', 'Value'])
        ws.append(['Product ID', product.id])
        ws.append(['Reference', product.reference])
        ws.append(['Name', product.name])
        ws.append(['Unit', product.get_unit_display()])
        ws.append(['Current Quantity', float(product.current_quantity)])
        ws.append(['Status', 'Active' if product.is_active else 'Inactive'])
        ws.append([])
        
        # Stock Movements
        movements = StockMovement.objects.filter(
            product=product,
            date__range=[start_date, end_date]
        ).order_by('-date')
        
        total_in = movements.filter(movement_type=StockMovement.MovementType.IN).aggregate(total=Sum('quantity'))['total'] or Decimal('0.00')
        total_out = movements.filter(movement_type=StockMovement.MovementType.OUT).aggregate(total=Sum('quantity'))['total'] or Decimal('0.00')
        
        ws.append(['STOCK SUMMARY'])
        ws.append(['Total IN', float(total_in)])
        ws.append(['Total OUT', float(total_out)])
        ws.append(['Net Movement', float(total_in - total_out)])
        ws.append([])
        
        ws.append(['ALL MOVEMENTS'])
        ws.append(['Date', 'Type', 'Quantity', 'Price (DA)', 'Total Value (DA)', 'Order', 'Created By'])
        
        for mov in movements:
            ws.append([
                mov.date.strftime('%Y-%m-%d %H:%M'),
                mov.get_movement_type_display(),
                float(mov.quantity),
                float(mov.price),
                float(mov.quantity) * float(mov.price),
                mov.order.order_number if mov.order else '',
                mov.created_by.username if mov.created_by else ''
            ])
        
        self.apply_header_styling(ws, title, subtitle)
    
    def _add_clients_detailed_sheet(self, wb, filters):
        ws = wb.create_sheet("Clients - Detailed")
        
        clients = Client.objects.all().prefetch_related('orders')
        
        # Apply filter
        if filters['client_id']:
            clients = clients.filter(id=filters['client_id'])
        
        title = "DETAILED CLIENTS REPORT"
        subtitle = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')} | Total Records: {clients.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Name', 'Phone', 'Address', 'Client Type', 'Credit Balance (DA)',
            'Status', 'Total Orders', 'Total Orders Value (DA)', 'Total Paid (DA)',
            'Outstanding Balance (DA)', 'Last Order Date', 'Last Order Number', 'Notes'
        ]
        ws.append(headers)
        
        for client in clients:
            orders = client.orders.all()
            total_orders = orders.count()
            total_value = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
            
            # Calculate total paid by summing CLIENT_PAYMENT inputs for this client's orders
            total_paid = Input.objects.filter(
                order__client=client,
                type=Input.Type.CLIENT_PAYMENT
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            outstanding = total_value - total_paid
            last_order = orders.order_by('-order_date').first()
            
            ws.append([
                client.name,
                client.phone,
                client.address,
                client.get_client_type_display(),
                float(client.credit_balance),
                'Active' if client.is_active else 'Inactive',
                total_orders,
                float(total_value),
                float(total_paid),
                float(outstanding),
                last_order.order_date.strftime('%Y-%m-%d') if last_order else 'No Orders Yet',
                last_order.order_number if last_order else '',
                client.notes
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
    
    def _add_suppliers_detailed_sheet(self, wb, filters):
        ws = wb.create_sheet("Suppliers - Detailed")
        
        suppliers = Supplier.objects.all()
        
        # Apply filter
        if filters['supplier_id']:
            suppliers = suppliers.filter(id=filters['supplier_id'])
        
        title = "DETAILED SUPPLIERS REPORT"
        subtitle = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')} | Total Records: {suppliers.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Name', 'Phone', 'Address', 'Status',
            'Total Expenses (DA)', 'Total Transactions', 'Last Transaction Date', 'Notes'
        ]
        ws.append(headers)
        
        for supplier in suppliers:
            outputs = Output.objects.filter(supplier=supplier)
            total_expenses = outputs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            transaction_count = outputs.count()
            last_output = outputs.order_by('-date').first()
            
            ws.append([
                supplier.name,
                supplier.phone or 'N/A',
                supplier.address or 'N/A',
                'Active' if supplier.is_active else 'Inactive',
                float(total_expenses),
                transaction_count,
                last_output.date.strftime('%Y-%m-%d') if last_output else 'No Transactions',
                supplier.notes
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
    
    def _add_products_detailed_sheet(self, wb, filters):
        ws = wb.create_sheet("Products - Detailed")
        
        products = Product.objects.all()
        
        # Apply filter
        if filters['product_id']:
            products = products.filter(id=filters['product_id'])
        
        title = "DETAILED PRODUCTS REPORT"
        subtitle = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')} | Total Records: {products.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Product Reference', 'Name', 'Unit', 'Current Quantity',
            'Total Stock IN', 'Total Stock OUT', 'Last Movement Date',
            'Last Movement Type', 'Status', 'Description'
        ]
        ws.append(headers)
        
        for product in products:
            in_movements = StockMovement.objects.filter(
                product=product, 
                movement_type=StockMovement.MovementType.IN
            )
            out_movements = StockMovement.objects.filter(
                product=product, 
                movement_type=StockMovement.MovementType.OUT
            )
            
            total_in = in_movements.aggregate(total=Sum('quantity'))['total'] or Decimal('0.00')
            total_out = out_movements.aggregate(total=Sum('quantity'))['total'] or Decimal('0.00')
            last_movement = StockMovement.objects.filter(product=product).order_by('-date').first()
            
            ws.append([
                product.reference,
                product.name,
                product.get_unit_display(),
                float(product.current_quantity),
                float(total_in),
                float(total_out),
                last_movement.date.strftime('%Y-%m-%d') if last_movement else 'No Movements',
                last_movement.get_movement_type_display() if last_movement else '',
                'Active' if product.is_active else 'Inactive',
                product.description
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
    
    def _add_users_detailed_sheet(self, wb):
        ws = wb.create_sheet("Users - Detailed")
        
        users = User.objects.all()
        
        title = "DETAILED USERS REPORT"
        subtitle = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')} | Total Records: {users.count()}"
        
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        
        headers = [
            'Username', 'Full Name', 'Role', 'Status',
            'Inputs Created', 'Outputs Created', 'Stock Movements Created', 'Last Login'
        ]
        ws.append(headers)
        
        for user in users:
            inputs_count = Input.objects.filter(created_by=user).count()
            outputs_count = Output.objects.filter(created_by=user).count()
            movements_count = StockMovement.objects.filter(created_by=user).count()
            
            ws.append([
                user.username,
                user.full_name,
                user.get_role_display(),
                'Active' if user.is_active else 'Inactive',
                inputs_count,
                outputs_count,
                movements_count,
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never'
            ])
        
        self.apply_header_styling(ws, title, subtitle)
        self.apply_table_styling(ws, header_row=4)
